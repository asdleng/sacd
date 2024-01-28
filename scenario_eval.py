from turtle import clear
import time
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from sacd.agent.mpc_new import MPC,MPC2,MPC3,MPC_des
from sacd.agent.sacd import SacdAgent
import gymnasium as gym
from sacd.agent.energy_cost import energy_cost
from sacd.agent.energy_cost_motor import energy_cost_motor
from sacd.agent.speed_generator import normalize_speed
import os
import torch
from torch.optim import Adam

def clear_directory(path):
    for item in os.listdir(path):
        item = os.path.join(path,item)
        if os.path.isfile(item):
            os.remove(item)
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
        
    sheet = workbook[sheet_name]
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)

# 0: use rule-based 1: DRL+MPC 2: MOBIL+IDM 3: Only Car-following MPC
# if use MOBIL+IDM model, you need also discommit line 127 and line 131 in action.py

# IMPORTANT!!!! # IMPORTANT!!!! # IMPORTANT!!!!

# DON'T FORGET TO DISCOMMENT LINE 228 in my_env.py and LINE 229: 'b_v_number-1', it should only restore for training
# Change the input of the LSTM model:
# LINE 51: input_size=1,
# LINE 75: speed_sequence = speed_sequence.view(speed_sequence.size(0), speed_sequence.size(1), 1)

MPC_protect = True
global_speed = True
behavior_flag = 1
scenario_num = 1
sur_spd = 10
seed = 1
spd_segs = []
cert_flag = False
cert_num = 0
DRL_times = []
MPC_times = []
s_t_segs = []
s_t_segs_tracking = []
vehicle_density = 1.0
upper = 20
lower = 0
# 3-stage or IDM ---- if use 'behavior_flag==4', then just ignore the M-IDM planned trajectory and always use the 3-stage one
if global_speed  == False:
    spd_file = "/home/i/sacd/ref_spd_data/idm_scenario"+str(scenario_num)+"/spd_seg_lin.txt"
    target_t_file = "/home/i/sacd/ref_spd_data/idm_scenario"+str(scenario_num)+"/target_times.txt"
else:
    spd_file = "/home/i/sacd/ref_spd_data/scenario"+str(scenario_num)+"/spd_seg_lin.txt"
    target_t_file = "/home/i/sacd/ref_spd_data/scenario"+str(scenario_num)+"/target_times.txt"
target_times = np.loadtxt(target_t_file)

with open(spd_file, "r") as file:
    for line in file:
        values = line.strip().split(",")
        spd_seg = np.array([float(value) for value in values])
        spd_segs.append(spd_seg)

def cal_s_from_tv(v_tv):
    s_t = []
    s_t.append(0)
    for i in range(len(v_tv)):
        ds = v_tv[i]
        s_t.append(s_t[-1]+ds)
    return s_t

def convert_sv_to_tv(v_sv):
    t_s = 1 / v_sv
    t_array = np.zeros(len(v_sv))
    for i in range(len(v_sv)-1):
        t_array[i+1] = t_array[i] + t_s[i]
    cur_x = 0
    v_tv = np.array([])
    t = 0
    while cur_x<len(v_sv)-1:
        v = np.interp(t,t_array,v_sv)
        v_tv = np.append(v_tv,v)
        t = t+1
        cur_x = np.interp(t,t_array,np.arange(len(t_array)))
    return v_tv



RENDER = True

policy_frequency = 1
env = gym.make("myenv",  render_mode='rgb_array')
state_dim = env.observation_space.shape[0] * \
            env.observation_space.shape[1]
act_dim = 3
env.configure(
            {
                "simulation_frequency": 10,
                "policy_frequency": 10,
                "duration": 500,
                "screen_width": 500,
                "screen_height": 200,
                "show_trajectories": False,
                "render_agent": False,
                "offscreen_rendering": False,
                "initial_lane_id": 1,
                "vehicles_density": vehicle_density,
                "ego_spacing":1.5
            })
env.config['initial_lane_id'] = 1

# load 模型
log_dir = "/home/i/sacd/logs/myenv"
sacdagent = SacdAgent(env,env,log_dir,has_speed=True,cuda=True,spd_type='lstm',method='sacd')
sacdagent.RENDER = True
#sacdagent.policy.load("/home/i/sacd/sacd/sacd_current_model/policy.pth")# 注意lower = 10
#sacdagent.policy.load("/home/i/sacd/scenario_eval.pth") # 注意lower = 8
#sacdagent.policy.load("/home/i/sacd/scenario_eval_lstm2.pth") # 注意lower = 10
#sacdagent.policy.load("/home/i/sacd/scenario_eval_lstm4.pth") # 注意lower = 0, 使用spd_seq
#sacdagent.policy.load("/home/i/sacd/scenario_eval_lstm5.pth") # 注意lower = 0, 使用eco_seq
#sacdagent.policy.load("/home/i/sacd/scenario_eval_lstm7.pth") # 注意lower = 0, 使用spd_seq
sacdagent.policy.load("/home/i/sacd/scenario_eval_lstm8.pth") # 注意lower = 0, 使用eco_seq
sacdagent.policy.eval()
total_return = 0
collision = 0
total_error = 0
total_stp = 0
total_delay = 0

# MPC_des
lane_change_num = 0
spd_segs_track = []
spd_segs_track_v = []
spd_segs_track_s = []
spd_segs_track_y = []
spd_segs_track_dis = []
spd_segs_track_ttc = []
veh_pos = [[] for _ in range(len(spd_segs))]
# 清空images文件夹
if behavior_flag==1:
    clear_directory("images")
for i in range(len(spd_segs)):
    print("====第",i+1,"/",len(spd_segs),"段====")
    spd_seg_track_s = np.array([])
    spd_seg_track_y = np.array([])
    spd_seg_track_v = np.array([])
    spd_seg_track_ttc = np.array([])
    spd_seg_track_dis = np.array([])
    env.set_ref_speed2(spd_segs[i],sur_spd)
    env.config['initial_lane_id'] = 1
    state = env.reset(seed = seed)
    change_flag  = False
    lane_change_count = 0
    target_lane = env.config["initial_lane_id"]
    no_solution_flag = False
    if (RENDER):
        env.config['offscreen_rendering'] = False
    seed+=1
    #state = state[0]
    #state = state.reshape([state_dim])
    state = sacdagent.make_state(env,state[0])
    episode_steps = 0
    episode_return = 0.0
    done, truncate = False, False
    last_current_lane = env.config['initial_lane_id']
    old_lane = env.config['initial_lane_id']
    stp = 0
    eposide_error = 0
    count = 0
    delay = 0
    last_action_is_change = False
    lane_change_finished = False
    start_record_trig = False
    while (not done and not truncate):
        veh_pos[i].append(env.vehicle.position)
        stp += 1
        if (RENDER == True):
            #plt.imshow(env.render())
            image = env.render()
            if start_record_trig:
                plt.imsave("images/"+str(i)+"_"+str(stp)+".jpg",image)
                record_count+=1
                if record_count>20:
                    record_count = 0
                    start_record_trig = False    
        spd_seg_track_v = np.append(spd_seg_track_v,env.vehicle.speed)
        spd_seg_track_s = np.append(spd_seg_track_s,env.vehicle.position[0]-env.start_position)
        spd_seg_track_y = np.append(spd_seg_track_y,env.vehicle.position[1])
        spd_seg_track_ttc = np.append(spd_seg_track_ttc,env.TTC)
        spd_seg_track_dis = np.append(spd_seg_track_dis,env.veh_dis)
        current_lane = np.clip(round(env.vehicle.position[1]/4),0,2)
        record_count = 0
        if behavior_flag == 0:  # Rule
            action_g,last_action_is_change,target_lane,change_flag,lane_change_count,no_solution_flag = MPC_des(env,last_action_is_change,count,target_lane,change_flag,lane_change_count,no_solution_flag,horizon = 10,dt = 0.4)
            if not isinstance(action_g,np.ndarray):
                action_g = np.array([0,0])
            
        elif behavior_flag == 1 or behavior_flag == 3: # DRL+MPC
            if count % round(env.config['policy_frequency']/policy_frequency) == 0:
                speed_seq = normalize_speed(env.get_speed_seq(100),upper,lower)
                eco_seq = env.get_eco_seq()
                t1 = time.time()
                #action_d = sacdagent.exploit(state,speed_seq)
                action_d = sacdagent.exploit(state,eco_seq)
                t2 = time.time()
                DRL_times.append(t2-t1)
                if action_d != 1:
                    start_record_trig = True
                if behavior_flag==3:
                    action_d = 1
                if action_d !=1:
                    ## 标记下换道前的车道
                    old_lane = current_lane
                if current_lane!=old_lane:
                    lane_change_finished = True
                else:
                    lane_change_finished = False
                # 换道完成，强制直行
                if last_action_is_change and lane_change_finished:
                    action_d = 1
                    last_action_is_change = False
                if action_d != 1:
                    last_action_is_change = True
            action_g, reward_no_solution, no_solution_done,delta_t,cert = sacdagent.get_MPC_actions(
                env,action_d, train=False, horizon=10, eval_MPC_protect=MPC_protect,dt=0.4)
            MPC_times.append(delta_t)
            if cert and not cert_flag:
                cert_num+=1
                cert_flag = True
            elif not cert and cert_flag:
                cert_flag = False

        elif behavior_flag == 2:
            dx = (env.get_t_x(env.time)-env.vehicle.position[0])
            K = 1
            tar_spd = env.get_ref_speed()[0]+dx*K
            env.vehicle.target_speed = tar_spd
            action_g = None

        next_state, reward, done, truncate, info = env.step(action_g)
        if last_current_lane != current_lane:
            lane_change_num+=1
            last_current_lane = current_lane
        if count % round(env.config['policy_frequency']/policy_frequency) == 0:
            episode_steps += 1
            episode_return += reward

        if behavior_flag == 1:
            last_d = action_d
            #next_state = np.array(next_state).reshape([state_dim])
            next_state = sacdagent.make_state(env,np.array(next_state))
            state = next_state
        error = abs(info['speed']-env.get_ref_speed()[0])
        eposide_error += error
        count += 1
        if done:
            collision+=1
    spd_segs_track_v.append(spd_seg_track_v)
    spd_segs_track_s.append(spd_seg_track_s)
    spd_segs_track_y.append(spd_seg_track_y)
    spd_segs_track_ttc.append(spd_seg_track_ttc)
    spd_segs_track_dis.append(spd_seg_track_dis)
    total_return += episode_return
    total_error += eposide_error
    total_stp += stp
    delay = max(env.time - target_times[i],0)
    total_delay+=abs(delay)
    print("段误差为：", round(eposide_error/stp,2))
    print("段奖励为：",round(episode_return,2))
    print("段终端延误为：",round(delay,2))

env.close()
if behavior_flag ==0:
    print("====使用的是Rule+MPC====")
elif behavior_flag == 1:
    print("====使用的是RL====")
elif behavior_flag == 2:
    print("====使用的是IDM+MOBIL====")
elif behavior_flag == 3:
    print("====使用的是Car-following MPC====")
print("总奖励为：", round(total_return,2))
print("总碰撞次数为：", collision)
print("平均误差为：", round(total_error/total_stp,2))
print("平均终端延误为：",round(total_delay/len(spd_segs),2))
print("换道次数为：",lane_change_num)
print("DRL耗时为：",np.mean(np.array(DRL_times)))
print("MPC耗时为：",np.mean(np.array(MPC_times)))
print("cert次数为：",cert_num)

book_name_xlsx = 'scenario_data.xlsx'
sheet_name_xlsx = 'scenario_'+str(scenario_num)

if behavior_flag == 0:
    sheet_name_xlsx = sheet_name_xlsx+'_Rule+MPC'
elif behavior_flag == 1:
    sheet_name_xlsx = sheet_name_xlsx+'_RL'
value = [["总奖励",str(round(total_return,2))],
                ["平均误差",str(round(total_error/total_stp,2))],
                [ "平均终端延误", str(round(total_delay/len(spd_segs),2))],
                [ "换道次数", lane_change_num],]

write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value)


total_energy_target = 0
total_energy_track = 0
for i in range(len(spd_segs)):
    spd_seg_vt = convert_sv_to_tv(spd_segs[i])
    s_t = cal_s_from_tv(spd_seg_vt)
    s_t_segs.append(s_t)
    acc_seg = np.diff(spd_seg_vt)
    spd_seg_vt = spd_seg_vt[:-1]
    energy_seg = energy_cost_motor(spd_seg_vt,acc_seg)
    total_energy_target+=energy_seg

    spd_seg_track = np.interp(np.arange(len(spd_segs[i])),spd_segs_track_s[i],spd_segs_track_v[i])
    spd_segs_track.append(spd_seg_track)
    spd_seg_vt_track = convert_sv_to_tv(spd_seg_track)
    s_t_track = cal_s_from_tv(spd_seg_vt_track)
    s_t_segs_tracking.append(s_t_track)
    acc_seg_track = np.diff(spd_seg_vt_track)
    spd_seg_vt_track = spd_seg_vt_track[:-1]
    energy_seg_track = energy_cost_motor(spd_seg_vt_track,acc_seg_track)
    total_energy_track+=energy_seg_track

print("Tracking能耗/Target能耗为：",total_energy_track,",",total_energy_target)
sub_row = int(np.sqrt(len(spd_segs))+1)
sub_col = int(np.sqrt(len(spd_segs))+1)

fig, axes = plt.subplots(sub_row,sub_col)
fig2, axes2 = plt.subplots(sub_row,sub_col)
k = 0

if global_speed == False:
    behavior_flag = 4

for i in range(sub_row):
    if(k>=len(spd_segs)):
        break
    for j in range(sub_col):
        if(k>=len(spd_segs)):
            break
        axes[i,j].plot(spd_segs[k],color='red')
        axes[i,j].plot(spd_segs_track[k],color='blue')
        axes[i, j].set_ylim(0, 20)
        axes[i, j].set_ylabel('m/s')
        file_path = "scenario_eval_traj/"+str(scenario_num)+"_"+str(k)+"_ref.txt"
        with open(file_path, 'w') as file:
            # Write each double number to the file
            for number in spd_segs[k]:
                file.write(str(number) + '\n')
        file_path = "scenario_eval_traj/"+str(scenario_num)+"_"+str(k)+"_trk_"+str(behavior_flag)+".txt"
        with open(file_path, 'w') as file:
            # Write each double number to the file
            for number in spd_segs_track[k]:
                file.write(str(number) + '\n')
        file_path = "scenario_eval_traj/"+str(scenario_num)+"_"+str(k)+"_xy_"+str(behavior_flag)+".txt"
        with open(file_path, 'w') as file:
            # Write each double number to the file
            for m in range(len(spd_segs_track_s[k])):
                str_to_write = str(spd_segs_track_s[k][m]) + "," + str(spd_segs_track_y[k][m])
                file.write(str_to_write + '\n')
        file_path = "scenario_eval_traj/"+str(scenario_num)+"_"+str(k)+"_ttc_"+str(behavior_flag)+".txt"
        with open(file_path, 'w') as file:
            # Write each double number to the file
            for number in spd_segs_track_ttc[k]:
                file.write(str(number) + '\n')
        file_path = "scenario_eval_traj/"+str(scenario_num)+"_"+str(k)+"_dis_"+str(behavior_flag)+".txt"
        with open(file_path, 'w') as file:
            # Write each double number to the file
            for number in spd_segs_track_dis[k]:
                file.write(str(number) + '\n')
        axes2[i,j].plot(s_t_segs[k],color='red')
        axes2[i,j].plot(s_t_segs_tracking[k],color='blue')
        axes2[i, j].set_ylabel('m')
        k = k+1
plt.tight_layout()
plt.show()